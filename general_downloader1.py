import openpyxl
import os
import time
from selenium import webdriver
import openpyxl
import io
import pandas as pd
from datetime import datetime, timedelta
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
import re


def remove_special_chars(string):
    pattern = r'[^\w\s]'
    return re.sub(pattern, '', string)


def lifefinance_scrap(href, dict_for_traders):
    site_name = 'litefinance'
    df_for_trader = pd.DataFrame(dict_for_traders)
    print(f'Перехожу по ссылке трейдера: {href.value}\n')
    options = webdriver.ChromeOptions()
    options.add_argument('chromedriver_binary.chromedriver_filename')
    # options.add_argument('headless')
    options.add_argument("window-size=1920,1080")
    driver = webdriver.Chrome(options=options)
    driver.maximize_window()
    driver.get(href.value)
    print(f'Успешно перешел по ссылке {href.value}\n')
    WebDriverWait(driver, 600).until(
        ec.presence_of_element_located(
            ("xpath", fr'//div[@class = "page_header_part traders_body"]//h2'))
    )
    name = remove_special_chars(
        driver.find_element(
            "xpath",
            fr'//div[@class = "page_header_part"]//h2').text
    )
    print(f'Имя трейдера = {name}\n')
    excel_name = fr'{bd_dir}\{site_name}\excel\{name}.xlsx'
    htm_name = fr'{bd_dir}\{site_name}\htm\{name}.htm'
    for o in (range(2, 10)):
        time.sleep(2)
        count = 0
        while count == 0:
            count = len(driver.find_elements("xpath",
                                             fr'//div[@class = "content_row"]'))
        print(f'Начинаю обработку {count} записей на странице {o - 1}\n')
        for l in list(range(count - 49, count + 1)):
            currency = driver.find_element("xpath",
                                           fr'(//div[@class = "content_row"])[{l}]/descendant::a[2]').text
            if currency is not None:
                date_close = driver.find_element("xpath",
                                                 fr'(//div[@class = "content_row"])[{l}]'
                                                 fr'/descendant::div[@class = "content_col"][3]').text
                date_close = datetime.strptime(date_close,
                                               '%d.%m.%Y %H:%M:%S') + timedelta(
                    hours=-1)
                date_open = driver.find_element("xpath",
                                                fr'(//div[@class = "content_row"])[{l}]'
                                                fr'/descendant::div[@class = "content_col"][2]').text
                date_open = datetime.strptime(date_open,
                                              '%d.%m.%Y %H:%M:%S') + timedelta(
                    hours=-1)
                type_of_trade = driver.find_element("xpath",
                                                    fr'(//div[@class = "content_row"])[{l}]'
                                                    fr'/descendant::div[@class = "content_col"][4]'
                                                    ).text.lower()
                if type_of_trade == 'покупка':
                    type_of_trade = 'buy'
                else:
                    type_of_trade = 'sell'
                obj = driver.find_element("xpath",
                                          fr'(//div[@class = "content_row"])[{l}]'
                                          fr'/descendant::div[@class = "content_col"][5]'
                                          ).text.replace(".", ",")
                currency = driver.find_element("xpath",
                                               fr'(//div[@class = "content_row"])[{l}]/descendant::a[2]') \
                    .text.replace("XAUUSD", "GOLD")
                price_open = driver.find_element("xpath",
                                                 fr'(//div[@class = "content_row"])[{l}]'
                                                 fr'/descendant::div[@class = "content_col"][6]'
                                                 ).text.replace(" ", "")
                price_close = driver.find_element("xpath",
                                                  fr'(//div[@class = "content_row"])[{l}]'
                                                  fr'/descendant::div[@class = "content_col"][7]'
                                                  ).text.replace(" ", "")
                points = driver.find_element("xpath",
                                             fr'(//div[@class = "content_row"])[{l}]'
                                             fr'/descendant::div[@class = "content_col"][8]'
                                             ).text.replace(".", ",")
                df_for_trader.loc[len(df_for_trader.index)] = [
                    obj,
                    currency,
                    type_of_trade,
                    date_open.strftime('%Y.%m.%d %H:%M'),
                    price_open,
                    date_close.strftime('%Y.%m.%d %H:%M'),
                    price_close,
                    points,
                ]
        driver.execute_script("arguments[0].scrollIntoView();",
                              driver.find_element("xpath",
                                                  fr'(//div[@class = "content_row"])[{count}]'))
    driver.quit()
    df_for_trader.to_excel(excel_name, sheet_name='Sheet1', index=False)
    # Дальнейший код нужен для красивого форматирования колонок в excel
    wb = openpyxl.load_workbook(excel_name)
    ws = wb.active
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            finally:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width
    wb.save(excel_name)
    print(f'✅ Успешно сформировал excel с сигналами трейдера {name}\n')
    text = ''
    for index, row in df_for_trader.iterrows():
        text = text + \
               '<tr align = right>' \
               fr'<td>{row["Объем"]}</td>' \
               fr'<td nowrap>{row["Время Открытия"]}</td>' \
               fr'<td>{row["Тип сделки"]}</td>' \
               fr'<td class=mspt>0</td>' \
               fr'<td>{row["Валютная пара"]}</td>' \
               fr'<td style="mso-number-format:0\.00000;">{row["Цена Открытия"]}</td>' \
               fr'<td style="mso-number-format:0\.00000;">{row["Объем"]}</td>' \
               fr'<td style="mso-number-format:0\.00000;">0</td>' \
               fr'<td class=msdate nowrap>{row["Время Закрытия"]}</td>' \
               fr'<td style="mso-number-format:0\.00000;">{row["Цена Закрытия"]}</td>' \
               fr'<td class=mspt>0</td>' \
               fr'<td class=mspt>0</td>' \
               fr'<td class=mspt>0</td>' \
               fr'<td class=mspt>0</td>' \
               '</tr>'
    with io.open(fr'{current_dir}\resources\template1.htm', 'r',
                 encoding='utf-8') as f:
        html_string = f.read()
    htm = html_string.replace("SSSSS", text)
    with io.open(htm_name, 'w', encoding='utf-8') as file:
        file.write(htm)
    print(f'✅ Успешно сформировал htm с сигналами трейдера {name}\n')


def forex4you_scrap(href, months_in_numbers, dict_for_traders):
    site_name = 'forex4you'
    df_for_trader = pd.DataFrame(dict_for_traders)
    print(f'Перехожу по ссылке трейдера: {href.value}\n')
    options = webdriver.ChromeOptions()
    # options.add_argument('headless')
    options.add_argument("window-size=1920,1080")
    driver = webdriver.Chrome(options=options)
    driver.maximize_window()
    driver.get(href.value)
    print(f'Успешно перешел по ссылке {href.value}\n')
    WebDriverWait(driver, 600).until(
        ec.presence_of_element_located(
            ("xpath",
             fr'//span[@data-ng-bind= "::$headerCtrl.leader.displayName"]'))
    )
    name = remove_special_chars(
        driver.find_element(
            "xpath",
            fr'//span[@data-ng-bind= "::$headerCtrl.leader.displayName"]'
        ).text
    )
    print(f'Имя трейдера = {name}\n')
    excel_name = fr'{bd_dir}\{site_name}\excel\{name}.xlsx'
    htm_name = fr'{bd_dir}\{site_name}\htm\{name}.htm'
    driver.find_element("xpath",
                        fr'//label[contains(text(), "Весь период")]'
                        ).click()
    for o in (range(2, 10)):
        time.sleep(2)
        count = 0
        while count == 0:
            count = len(driver.find_elements("xpath",
                                             fr'//tbody//tr[@data-ng-repeat = '
                                             fr'"trade in $fxGrid.$data track by trade.id"]'
                                             fr'//td[@data-ng-bind="::trade.symbol"]'))
        print(f'Начинаю обработку {count - 10} записей на странице {o - 1}\n')
        for l in list(range(1, count - 9)):
            currency = driver.find_element("xpath",
                                           fr'(//td[@data-ng-bind="::trade.symbol"])[{l}]').text
            if currency is not None:
                date_close = driver.find_element("xpath",
                                                 fr'(//td[@data-ng-bind="::trade.symbol"])[{l}]'
                                                 '//../preceding-sibling::td[1]').text
                for i in months_in_numbers:
                    date_close = date_close.replace(i, months_in_numbers[i])
                date_close = datetime.strptime(date_close,
                                               '%d %m %Y г., %H:%M:%S')
                date_open = driver.find_element("xpath",
                                                fr'(//td[@data-ng-bind="::trade.symbol"])[{l}]'
                                                fr'//../preceding-sibling::td[2]').text
                for i in months_in_numbers:
                    date_open = date_open.replace(i, months_in_numbers[i])
                date_open = datetime.strptime(date_open,
                                              '%d %m %Y г., %H:%M:%S')
                type_of_trade = driver.find_element("xpath",
                                                    fr'(//td[@data-ng-bind="::trade.symbol"])[{l}]'
                                                    fr'//../following-sibling::td[1]').text.lower()
                obj = driver.find_element("xpath",
                                          fr'(//td[@data-ng-bind="::trade.symbol"])[{l}]'
                                          fr'//../preceding-sibling::td[3]').text.replace(
                    ".", ",")
                currency = driver.find_element("xpath",
                                               fr'(//td[@data-ng-bind="::trade.symbol"])[{l}]').text \
                    .replace("XAUUSD", "GOLD")
                price_open = driver.find_element("xpath",
                                                 fr'(//td[@data-ng-bind="::trade.symbol"])[{l}]'
                                                 fr'//../following-sibling::td[2]').text.replace(
                    " ",
                    "")
                price_close = driver.find_element("xpath",
                                                  fr'(//td[@data-ng-bind="::trade.symbol"])[{l}]'
                                                  fr'//../following-sibling::td[3]').text.replace(
                    " ",
                    "")
                points = driver.find_element("xpath",
                                             fr'(//td[@data-ng-bind="::trade.symbol"])[{l}]'
                                             fr'//../following-sibling::td[4]').text.replace(
                    ".", ",")

                init_dict['df_for_trader'] = pd.concat([pd.DataFrame([[
                    obj,
                    currency,
                    type_of_trade,
                    date_open.strftime('%Y.%m.%d %H:%M'),
                    price_open,
                    date_close.strftime('%Y.%m.%d %H:%M'),
                    price_close,
                    points,
                ]], columns=init_dict['df_for_trader'].columns),
                    init_dict['df_for_trader']], ignore_index=True)



                df_for_trader.loc[len(df_for_trader.index)] = [
                    obj,
                    currency,
                    type_of_trade,
                    date_open.strftime('%Y.%m.%d %H:%M'),
                    price_open,
                    date_close.strftime('%Y.%m.%d %H:%M'),
                    price_close,
                    points,
                ]
        driver.find_element("xpath",
                            fr'(//a[@data-fx-grid-set-page="$fxGridPaginator.getNextPage()"])[1]').click()
    driver.quit()
    df_for_trader.to_excel(excel_name, sheet_name='Sheet1', index=False)
    # Дальнейший код нужен для красивого форматирования колонок в excel
    wb = openpyxl.load_workbook(excel_name)
    ws = wb.active
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            finally:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width
    wb.save(excel_name)
    print(f'✅ Успешно сформировал excel с сигналами трейдера {name}\n')
    text = ''
    for index, row in df_for_trader.iterrows():
        text = text + \
               '<tr align = right>' \
               fr'<td>{row["Объем"]}</td>' \
               fr'<td nowrap>{row["Время Открытия"]}</td>' \
               fr'<td>{row["Тип сделки"]}</td>' \
               fr'<td class=mspt>0</td>' \
               fr'<td>{row["Валютная пара"]}</td>' \
               fr'<td style="mso-number-format:0\.00000;">{row["Цена Открытия"]}</td>' \
               fr'<td style="mso-number-format:0\.00000;">{row["Объем"]}</td>' \
               fr'<td style="mso-number-format:0\.00000;">0</td>' \
               fr'<td class=msdate nowrap>{row["Время Закрытия"]}</td>' \
               fr'<td style="mso-number-format:0\.00000;">{row["Цена Закрытия"]}</td>' \
               fr'<td class=mspt>0</td>' \
               fr'<td class=mspt>0</td>' \
               fr'<td class=mspt>0</td>' \
               fr'<td class=mspt>0</td>' \
               '</tr>'
    with io.open(fr'{current_dir}\resources\template1.htm', 'r',
                 encoding='utf-8') as f:
        html_string = f.read()
    htm = html_string.replace("SSSSS", text)
    with io.open(htm_name, 'w', encoding='utf-8') as file:
        file.write(htm)
    print(f'✅ Успешно сформировал htm с сигналами трейдера {name}\n')


def make_hrefs_list(hrefs_file):
    input_excel = openpyxl.load_workbook(hrefs_file)
    sheet = input_excel['Лист1']
    list_of_input_hrefs = sheet['A']
    return list_of_input_hrefs


months_in_numbers = {"янв.": "01",
                     "февр.": "02",
                     "мар.": "03",
                     "апр.": "04",
                     "мая": "05",
                     "июня": "06",
                     "июля": "07",
                     "авг.": "08",
                     "сент.": "09",
                     "окт.": "10",
                     "нояб.": "11",
                     "дек.": "12",
                     }
dict_for_traders = {'Объем': [],
                    'Валютная пара': [],
                    'Тип сделки': [],
                    'Время Открытия': [],
                    'Цена Открытия': [],
                    'Время Закрытия': [],
                    'Цена Закрытия': [],
                    'Прибыль': [],
                    }
current_dir = os.path.dirname(os.path.abspath(__file__))
bd_dir = current_dir + r'\resources\БАЗА ДАННЫХ'
input_lists = [
    make_hrefs_list(bd_dir + r'\litefinance hrefs.xlsx'),
    make_hrefs_list(bd_dir + r'\forex4you hrefs.xlsx')
]

for site in input_lists:
for href in site:
    if href is None:
        continue
    elif 'forex4you' in href.value:
        forex4you_scrap(href, months_in_numbers, dict_for_traders)
    elif 'litefinance' in href.value:
        lifefinance_scrap(href, dict_for_traders)
